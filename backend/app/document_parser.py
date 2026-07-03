from pathlib import Path


SUPPORTED_EXTENSIONS = {".txt", ".md", ".pdf", ".docx", ".xlsx"}


def extract_text(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix not in SUPPORTED_EXTENSIONS:
        raise ValueError(f"不支持的文件类型：{suffix or '未知'}")
    if suffix in {".txt", ".md"}:
        return path.read_text(encoding="utf-8", errors="replace")
    if suffix == ".pdf":
        from pypdf import PdfReader

        return "\n".join(page.extract_text() or "" for page in PdfReader(path).pages)
    if suffix == ".docx":
        from docx import Document

        document = Document(path)
        return "\n".join(paragraph.text for paragraph in document.paragraphs if paragraph.text.strip())
    if suffix == ".xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        lines = []
        for sheet in workbook.worksheets:
            lines.append(f"工作表：{sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                values = [str(value) for value in row if value is not None]
                if values:
                    lines.append("\t".join(values))
        return "\n".join(lines)
    raise ValueError("无法解析文件")
